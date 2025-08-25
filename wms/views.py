from django.shortcuts import render, redirect, get_object_or_404
from .models import Item, Subdepartement, Transaction, TransactionItem, Location
from .forms import ItemForm, SubdepartementForm, TransactionForm, TransactionItemForm, TransactionItemFormSet, LocationForm
from django.forms import inlineformset_factory
from django.db.models import Sum, F, Case, When, IntegerField
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Q 
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView


class CustomLoginView(LoginView):
    template_name = 'wms/signin.html'

    def form_valid(self, form):
        # login user dulu
        response = super().form_valid(form)

        #cek apakah remme di centang
        if not self.request.POST.get("remember_me"):
            # session berlaku sampai browser close
            self.request.session.set_expiry(0)
        else:
            # misal sampai 30hari
            self.request.session.set_expiry(60 * 60 * 24 * 7)

        return response

@login_required
def logout_view(request):
    logout(request)
    return redirect("wms:logout")

@login_required
def profile(request):
    return render(request, "wms/profile.html", {'user':request.user})

@login_required
def dashboard(request):
    total_item = Item.objects.count()
    total_location = Location.objects.count()
    total_in = Transaction.objects.filter(transaction_type='IN').count()
    total_out = Transaction.objects.filter(transaction_type='OUT').count()

    # grafik transaksi IN vs OUT
    in_count = Transaction.objects.filter(transaction_type='IN').count()
    out_count = Transaction.objects.filter(transaction_type='OUT').count()

    top_outgoing_item = (
        TransactionItem.objects.filter(transaction__transaction_type="OUT").values("item__name").annotate(total_out=Sum("qty")).order_by("-total_out")[:5]
    )

    subdept_data = (
        Transaction.objects.filter(transaction_type="OUT").values("subdepartement__name").annotate(total_qty=Sum("items__qty")).order_by("subdepartement__name", "-total_qty")
    )

    subdept_labels = [d["subdepartement__name"] for d in subdept_data]
    subdept_values = [d["total_qty"] for d in subdept_data]


    # stok kritis
    items = Item.objects.annotate(
        total_in=Sum(
            Case(
                When(transaction_items__transaction__transaction_type='IN', then=F('transaction_items__qty')),
                output_field=IntegerField()
            )
        ),
        total_out=Sum(
            Case(
                When(transaction_items__transaction__transaction_type='OUT', then=F('transaction_items__qty')),
                output_field=IntegerField()
            )
        ),
    ).annotate(
        calc_stock=F('total_in') - F('total_out')
    )
    
    low_stock = items.filter(calc_stock__lte=5)[:5]

    # transaksi terbaru
    latest_in = Transaction.objects.filter(transaction_type="IN").order_by("-date")[:5]
    latest_out = Transaction.objects.filter(transaction_type="OUT").order_by("-date")[:5]
    recent_transactions = Transaction.objects.order_by('-date')[:5]

    context = {
        'title':"Dashboard",
        'total_item': total_item,
        'total_location': total_location,
        'total_in': total_in,
        'total_out': total_out,
        'in_count': in_count,
        'out_count': out_count,
        'low_stock': low_stock,
        'recent_transactions': recent_transactions,
        'latest_in': latest_in,
        'latest_out': latest_out,
        'top_outgoing_item': top_outgoing_item,
        'subdept_label': subdept_labels,
        'subdept_value': subdept_values,
    }
    return render(request, 'wms/dashboard.html', context)

@login_required
def location_list(request):
    locations = Location.objects.all().order_by('name')

    paginator = Paginator(locations, 10)
    page_number = request.GET.get('page')
    location = paginator.get_page(page_number)

    context = {
        'title':'List Location',
        'locations':location,
    }
    return render(request, 'wms/location/location_list.html', context)

@login_required
def add_location(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('wms:location_list')
    else:
        form = LocationForm()
    context = {
        'title':'Form Location',
        'form':form
    }

    return render(request, 'wms/location/location_form.html', context)

@login_required
def update_location(request, pk):
    location = get_object_or_404(Location, pk=pk)

    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            return redirect('wms:location_list')
    else:
        form = LocationForm(instance=location)

    context = {
        'title': 'Update Location',
        'form': form
    }
    return render(request, 'wms/location/location_form.html', context)

@login_required
def delete_location(request, pk):
    location = get_object_or_404(Location, pk=pk)

    if request.method == 'POST':
        location.delete()
    return redirect('wms:location_list')

@login_required
def export_locationToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Location"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Location"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Name']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Location.objects.all(), start=1):
        ws.append([
            idx,
            s.name,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Location.xlsx'
    wb.save(response)
    return response

@login_required
def print_locationToPdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition']='attachment; filename="location.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # judul
    title = Paragraph("Location Report", styles['Title'])
    elements.append(title)

    # date
    date_str = datetime.now().strftime("%d %B %Y, %H:%M")
    elements.append(Paragraph(f"Date: {date_str}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # ambil data dari model
    location = Location.objects.all().order_by("name")

    # header tabel
    data = [['No', 'Name']]

    # isi table
    for i, location in enumerate(location, start=1):
        data.append([
            i, 
            location.name,
        ])

    # buat tabel
    table = Table(data, colWidths=[40,200])
    
    # style table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

@login_required
def filter_location(request):
    query = request.GET.get('q','').strip()
    print("receive query:", query)

    if query:
        locations = Location.objects.filter(
            Q(name__icontains=query)
        )
    else:
        locations = Location.objects.all()
    
    location_list = []
    for location in locations:

        location_list.append({
            'id':location.id,
            'name':location.name,
        })

    return JsonResponse({'location':location_list})

@login_required
def item_list(request):
    item_lists = Item.objects.all().order_by('name')

    paginator = Paginator(item_lists, 10)
    page_number = request.GET.get('page')
    items = paginator.get_page(page_number)

    context = {
        'title':'List Item',
        'items':items,
    }
    return render(request, 'wms/item/item_list.html', context)

@login_required
def add_item(request):
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('wms:item_list')
    else:
        form = ItemForm()
    
    context = {
        'title':'Form Item',
        'form':form,
    }
    return render(request, 'wms/item/item_form.html', context)

@login_required
def update_item(request, pk):
    item = get_object_or_404(Item, pk=pk)

    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('wms:item_list')
    else:
        form = ItemForm(instance=item)

    context = {
        'title': 'Update Item',
        'form': form
    }
    return render(request, 'wms/item/item_form.html', context)

@login_required
def delete_item(request, pk):
    item = get_object_or_404(Item, pk=pk)

    if request.method == 'POST':
        item.delete()
    return redirect('wms:item_list')

@login_required
def export_itemToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Item"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Item"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Name', 'Unit', 'Location', 'Picture', 'Stock']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Item.objects.all(), start=1):
        ws.append([
            idx,
            s.name,
            s.unit,
            s.location.name,
            str(s.picture.url if s.picture else ""),  # konversi ke string
            s.stock,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Item.xlsx'
    wb.save(response)
    return response

@login_required
def filter_item(request):
    query = request.GET.get('q','').strip()
    print("receive query:", query)

    if query:
        items = Item.objects.filter(
            Q(name__icontains=query)|Q(location_name__icontains=query)
        )
    else:
        items = Item.objects.all()
    
    item_list = []
    for item in items:

        item_list.append({
            'id':item.id,
            'name':item.name,
            'unit':item.unit,
            'location':item.location.name,
            'picture':item.picture.url if item.picture else None,
        })

    return JsonResponse({'item':item_list})

@login_required
def print_itemToPdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition']='attachment; filename="item.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # judul
    title = Paragraph("Item Report", styles['Title'])
    elements.append(title)

    # date
    date_str = datetime.now().strftime("%d %B %Y, %H:%M")
    elements.append(Paragraph(f"Date: {date_str}", styles['Normal']))
    elements.append(Spacer(12, 12))

    # ambil data dari model
    item = Item.objects.all().order_by("name")

    # header tabel
    data = [[ 'No', 'Name', 'Unit', 'Location', 'Stock' ]]

    # isi table
    for i, items in enumerate(item, start=1):
        data.append([
            i, 
            items.name,
            items.unit,
            items.location,
            items.stock,
        ])

    # buat tabel
    table = Table(data, colWidths=[40, 200, 50, 80, 50])
    
    # style table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

@login_required
def subdepartement_list(request):
    subdepts = Subdepartement.objects.all().order_by('name')

    paginator = Paginator(subdepts, 10)
    page_number = request.GET.get('page')
    subdept = paginator.get_page(page_number)

    context = {
        'title':'List Sub-Departement',
        'subdepts':subdept,
    }
    return render(request, 'wms/subdept/subdept_list.html', context)

@login_required
def add_subdepartement(request):
    if request.method == 'POST':
        form = SubdepartementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('wms:subdept_list')
    else:
        form = SubdepartementForm()
    
    context = {
        'title':'Form Sub-Departement',
        'form':form,
    }
    return render(request, 'wms/subdept/subdept_form.html', context)

@login_required
def update_subdepartement(request, pk):
    subdept = get_object_or_404(Subdepartement, pk=pk)

    if request.method == 'POST':
        form = SubdepartementForm(request.POST, instance=subdept)
        if form.is_valid():
            form.save()
            return redirect('wms:subdept_list')
    else:
        form = SubdepartementForm(instance=subdept)

    context = {
        'title': 'Update Sub-Departement',
        'form': form
    }
    return render(request, 'wms/subdept/subdept_form.html', context)

@login_required
def delete_subdepartement(request, pk):
    subdept = get_object_or_404(Subdepartement, pk=pk)

    if request.method == 'POST':
        subdept.delete()
    return redirect('wms:subdept_list')

@login_required
def export_subdeptToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Sub-Departement"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Sub-Departement"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Name', 'Leader']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Subdepartement.objects.all(), start=1):
        ws.append([
            idx,
            s.name,
            s.leader,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Subdept.xlsx'
    wb.save(response)
    return response

@login_required
def filter_subdepartement(request):
    query = request.GET.get('q','').strip()
    print("receive query:", query)

    if query:
        subdepts = Subdepartement.objects.filter(
            Q(name__icontains=query)|Q(leader__icontains=query)
        )
    else:
        subdepts = Subdepartement.objects.all()
    
    subdepts_list = []
    for subdept in subdepts:

        subdepts_list.append({
            'id':subdept.id,
            'name':subdept.name,
            'leader':subdept.leader,
        })

    return JsonResponse({'subdept':subdepts_list})

@login_required
def print_subdeptToPdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition']='attachment; filename="Subdepartement.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        alignment=1, #0=left 1=center 2=right
        fontSize=16,
        spaceAfter=6,
    )
    dateStyle = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        alignment=1,
        fontsize=10,
        textColor=colors.grey,
        spaceAfter=20,
    )

    # judul
    title = Paragraph("Sub-Departement Report", title_style)
    elements.append(title)

    # date
    date_str = datetime.now().strftime("%d %B %Y, %H:%M")
    date_paragraph=Paragraph(f"Date: {date_str}", dateStyle)
    elements.append(date_paragraph)

    # spacer sebelum table
    elements.append(Spacer(1, 12))

    # ambil data dari model
    subdept = Subdepartement.objects.all().order_by("name")

    # header tabel
    data = [[ 'No', 'Name', 'Leader' ]]

    # isi table
    for i, subdept in enumerate(subdept, start=1):
        data.append([
            i, 
            subdept.name,
            subdept.leader,
        ])

    # buat tabel
    table = Table(data, colWidths=[40, 200, 150])
    
    # style table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

@login_required
def transaction_list(request):
    transaction_item = Transaction.objects.annotate(total_qty=Sum('items__qty')).order_by('-date')

    paginator = Paginator(transaction_item, 10)
    page_number = request.GET.get('page')
    transaction_item = paginator.get_page(page_number)

    context = {
         'title':"List Transaction",
         'transactions' : transaction_item
     }
    return render(request, 'wms/transaction/transaction_list.html', context)

@login_required
def add_transaction(request):
    TransactionItemFormSet = inlineformset_factory(
        Transaction, TransactionItem, form=TransactionItemForm, extra=1, can_delete=True
    )
    if request.method == 'POST':
        transaction_form = TransactionForm(request.POST)
        formset = TransactionItemFormSet(request.POST)

        if transaction_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    transaction_obj = transaction_form.save()
                    has_error =False
                    items = formset.save(commit=False)

                    for form in formset.forms:
                        item = form.save(commit=False)
                        if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                            continue

                        item = form.save(commit=False)
                        item.transaction = transaction_obj

                        if transaction_obj.transaction_type =="OUT":
                            if item.qty > item.item.stock:
                                form.add_error(
                                    "qty",
                                    f"Stock {item.item.name} is not enough! (ready stock : {item.item.stock})."
                                )
                                has_error = True
                                continue
                        try:
                            item.full_clean()
                        except ValidationError as e:
                            for msg in e.messages:
                                form.add_error(None, msg)

                        item.save()
                    
                    formset.save_m2m()

                    if has_error:
                        raise ValidationError("Ada error pada item formset")
                if not has_error:
                    messages.success(request, "Transaction succesfully created.")
                    return redirect('wms:transaction_list')
                
            except ValidationError:
                pass

    else:
        transaction_form = TransactionForm()
        formset = TransactionItemFormSet()

    context = {
        'title':'Form Transaction',
        'forms':transaction_form,
        'formset':formset
    }

    return render(request, 'wms/transaction/transaction_form.html', context)

@login_required
def export_transaksiToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Transaksi"

    # Judul besar di baris 1
    # ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    # ws['A1'] = "List Transaksi"
    # ws['A1'].font = Font(size=14, bold=True)
    # ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Type', 'Date', 'requested_by', 'received_by', 'Sub-departement', 'Item', 'Qty']
    ws.append(headers)

    # Data
    row_num = 2
    transactions = Transaction.objects.prefetch_related("items__item").all().order_by("-date")
    for idx, trx in enumerate(transactions, start=1):
        if trx.items.exists():
            for trx_item in trx.items.all():
                ws.append([
                    idx,
                    trx.transaction_type,
                    trx.date.strftime("%Y-%m-%d %H:%M"),
                    trx.requested_by,
                    trx.received_by,
                    trx.subdepartement.name,
                    trx_item.item.name,
                    trx_item.qty,
                    trx.note,
                ])
        else:
            ws.append([
                idx,
                trx.transaction_type,
                trx.date.strftime("%Y-%m-%d %H:%M"),
                trx.requested_by,
                trx.received_by,
                trx.subdepartement.name,
                "-",
                0,
                trx.note,
                ])
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Transaction.xlsx'
    wb.save(response)
    return response

@login_required
def print_transactionToPdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition']='attachment; filename="Transaction.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        alignment=1, #0=left 1=center 2=right
        fontSize=16,
        spaceAfter=5,
    )
    dateStyle = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        alignment=1,
        fontsize=10,
        textColor=colors.grey,
        spaceAfter=10,
    )

    # judul
    title = Paragraph("Transaction Report", title_style)
    elements.append(title)

    # date
    date_str = datetime.now().strftime("%d %B %Y, %H:%M")
    date_paragraph=Paragraph(f"Date: {date_str}", dateStyle)
    elements.append(date_paragraph)

    # spacer sebelum table
    elements.append(Spacer(1, 12))

    # ambil data dari model
    transactions = Transaction.objects.all().order_by("-date")

    # header tabel
    data = [[ 'No', 'Date', 'Type', 'Requested By', 'received_by', 'Sub-Dept', 'item', 'qty', 'note']]

    # isi table
    for idx, trx in enumerate(transactions, start=1):
        if trx.items.exists():
            # gabungkan semua item
            item_str = "\n".join([f"{trx_item.item.name} ({trx_item.qty})" for trx_item in trx.items.all()])
            total_qty = sum([trx_item.qty for trx_item in trx.items.all()])
        else:
            items_str = "-"
            total_qty = 0
        
        data.append([
            idx,
            trx.date.strftime("%Y-%m-%d %H:%M"),
            trx.transaction_type,
            trx.requested_by,
            trx.received_by,
            trx.subdepartement.name,
            item_str,
            total_qty,
            trx.note,
        ])
        
    # buat tabel
    table = Table(data, colWidths=[40, 90, 30, 100, 100, 80, 150, 40, 100])
    
    # style table
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

@login_required
def filter_transaction(request):
    query = request.GET.get('q','').strip()
    transaction_type=request.GET.get('transaction_type', '').strip()

    print("reseive query:", query, "transaction-type:", transaction_type)

    transactions = Transaction.objects.all()


    if query:
        transactions = transactions.filter(
            Q(transaction_type__iexact=query)
            |Q(date__icontains=query)
            |Q(requested_by__icontains=query)
            |Q(received_by__icontains=query)
            |Q(subdepartement__name__icontains=query)
            |Q(items__name__icontains=query)
        ).distinct()

    if transaction_type: #hanya filter saat ada pilihan
        transactions = transactions.filter(transaction_type=transaction_type)

    transaction_list = []
    for transaction in transactions:
        total_qty = sum(i.qty for i in transaction.items.all())  # <-- hitung manual
        transaction_list.append({
            'id':transaction.id,
            'transaction_type':transaction.transaction_type,
            'date':transaction.date.strftime("%Y-%m-%d"),
            'items':[
                {'name':item.item.name, 'qty':item.qty} for item in transaction.items.all()
                ],
            'total_qty':total_qty,
            'requested_by':transaction.requested_by,
            'received_by':transaction.received_by,
            'subdepartement':transaction.subdepartement.name if transaction.subdepartement else "",
            'note':transaction.note,
        })

    return JsonResponse({'transaction':transaction_list})